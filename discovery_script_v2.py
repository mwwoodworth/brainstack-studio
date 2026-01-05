import openpyxl
import os
import glob

# Path to estimates
ESTIMATES_DIR = "/home/matt-woodworth/Desktop/SOPS-ERP/Estimating/WCEstimates/"

# Find one xlsx file
files = glob.glob(os.path.join(ESTIMATES_DIR, "*.xlsx"))
sample_file = files[0]
print(f"Inspecting: {sample_file}")

try:
    wb = openpyxl.load_workbook(sample_file, data_only=True)
    if 'SUMMARY' in wb.sheetnames:
        ws = wb['SUMMARY']
        print(f"Sheet: SUMMARY")
        print("-" * 20)
        print("First 40 rows:")
        for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True)):
             cleaned_row = [cell for cell in row if cell is not None]
             if cleaned_row:
                 print(f"Row {i+1}: {cleaned_row}")
    else:
        print("SUMMARY sheet not found.")

except Exception as e:
    print(f"Error: {e}")
