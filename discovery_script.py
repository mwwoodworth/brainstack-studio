import openpyxl
import os
import glob

# Path to estimates
ESTIMATES_DIR = "/home/matt-woodworth/Desktop/SOPS-ERP/Estimating/WCEstimates/"

# Find one xlsx file
files = glob.glob(os.path.join(ESTIMATES_DIR, "*.xlsx"))
if not files:
    print("No .xlsx files found.")
    exit()

sample_file = files[0]
print(f"Inspecting: {sample_file}")

try:
    wb = openpyxl.load_workbook(sample_file, data_only=True)
    print("Sheet Names:", wb.sheetnames)
    
    # Inspect active sheet
    ws = wb.active
    print(f"Active Sheet: {ws.title}")
    
    print("-" * 20)
    print("First 20 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        # Filter None to clean output
        cleaned_row = [cell for cell in row if cell is not None]
        if cleaned_row:
             print(f"Row {i+1}: {cleaned_row}")

except Exception as e:
    print(f"Error: {e}")
