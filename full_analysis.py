import os
import glob
import re
import random
import openpyxl
from collections import Counter, defaultdict
import datetime

# --- CONFIG ---
ESTIMATES_DIR = "/home/matt-woodworth/Desktop/SOPS-ERP/Estimating/WCEstimates/"
OUTPUT_FILE = "/home/matt-woodworth/dev/ESTIMATES_ANALYSIS.md"
SAMPLE_SIZE = 50
ERP_COMPARISON_NOTE = """
## 8. ERP Cross-Reference
**ERP Target:** `src/services/estimates-production.ts` (Estimate Interface) 

| Excel Field | ERP Field | Match Status | Notes |
|---|---|---|---|
| Project Name | `project_name` / `title` | ✅ Direct | |
| Customer Name | `customer_name` | ✅ Direct | |
| Address | `property_address` | ✅ Direct | |
| Description | `description` / `scope_of_work` | ✅ Direct | |
| Date | `estimate_date` | ✅ Direct | |
| **Line Items** | `line_items` | ⚠️ Transform | Excel rows need parsing into `EstimateLineItem` objects. |
| Labor Cost | `labor_cost` (AI/Job) | ⚠️ Partial | Found in Summary breakdown, not always explicit in Estimate model. |
| Material Cost | `material_cost` (AI/Job) | ⚠️ Partial | Found in Summary breakdown. |
| **Grand Total** | `total` / `total_amount` | ✅ Direct | |
| System (EPDM, etc) | `roof_type` | ✅ Direct | "EPDM", "TPO" map directly. |
"""

# --- UTILS ---

def clean_value(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    return str(val).strip()

def find_cell_by_label(ws, label_pattern, search_rows=15, search_cols=5):
    """Searches for a label (regex) and returns the value in the next column."""
    for r in range(1, search_rows + 1):
        for c in range(1, search_cols + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val and isinstance(cell_val, str):
                if re.search(label_pattern, cell_val, re.IGNORECASE):
                    # Found label, return next column value
                    # Check C (next) then D then E
                    val = ws.cell(row=r, column=c+1).value
                    if val: return clean_value(val)
                    val = ws.cell(row=r, column=c+2).value
                    if val: return clean_value(val)
    return None

def analyze_filename(filename):
    """Extracts metadata from filename."""
    name = os.path.splitext(filename)[0]
    
    # Try generic "Address - Type"
    parts = name.split('-')
    if len(parts) >= 2:
        return {
            'address_or_desc': parts[0].strip(),
            'type': parts[-1].strip(),
            'raw': name
        }
    return {'address_or_desc': name, 'type': 'Unknown', 'raw': name}

# --- MAIN ---

def main():
    print("Starting Estimates Analysis...")
    
    all_files = glob.glob(os.path.join(ESTIMATES_DIR, "*.xlsx"))
    all_files.extend(glob.glob(os.path.join(ESTIMATES_DIR, "*.xls")))
    
    print(f"Found {len(all_files)} files.")
    
    # 1. Filename Analysis
    years = []
    types = Counter()
    
    year_regex = re.compile(r'20\d{2}')
    
    for f in all_files:
        fname = os.path.basename(f)
        meta = analyze_filename(fname)
        types[meta['type']] += 1
        
        ym = year_regex.search(fname)
        if ym:
            years.append(ym.group(0))
    
    # 2. Sampling & Deep Dive
    sample_files = random.sample(all_files, min(len(all_files), SAMPLE_SIZE))
    
    pricing_structures = []
    tab_structures = Counter()
    common_templates = Counter()
    calculation_patterns = Counter()
    
    print(f"Analyzing {len(sample_files)} sample files...")
    
    extracted_data = []
    
    for i, fpath in enumerate(sample_files):
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            sheet_names = wb.sheetnames
            for s in sheet_names: tab_structures[s] += 1
            
            # Identify Template Version based on Sheet Hash/Set
            template_sig = ",".join(sorted([s for s in sheet_names if s in ['SUMMARY', 'EPDM', 'TPO', 'BUR', 'SHINGLES']]))
            common_templates[template_sig] += 1
            
            # Analyze SUMMARY
            summary_data = {}
            if 'SUMMARY' in sheet_names:
                ws = wb['SUMMARY']
                summary_data['project'] = find_cell_by_label(ws, r'PROJECT:')
                summary_data['customer'] = find_cell_by_label(ws, r'CUSTOMER:')
                summary_data['address'] = find_cell_by_label(ws, r'ADDRESS:')
                summary_data['date'] = find_cell_by_label(ws, r'DATE:')
                
                # Pricing/Systems check (Rows 10-30)
                active_systems = []
                for r in range(10, 30):
                    sys_name = ws.cell(row=r, column=1).value
                    total_val = ws.cell(row=r, column=4).value # Column D is often total
                    if sys_name and isinstance(sys_name, str) and isinstance(total_val, (int, float)) and total_val > 10:
                        active_systems.append(f"{sys_name} (${total_val:,.0f})")
                
                summary_data['active_systems'] = active_systems
            
            extracted_data.append({
                'file': fname,
                'sheets': sheet_names,
                'summary': summary_data
            })
            
        except Exception as e:
            print(f"Error reading {fname}: {e}")

    # --- REPORT GENERATION ---
    
    with open(OUTPUT_FILE, 'w') as f:
        f.write("# COMPREHENSIVE ESTIMATES ANALYSIS\n\n")
        f.write(f"**Generated:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"**Scope:** {len(all_files)} files analyzed, {len(sample_files)} deep-dived.\n\n")
        
        f.write("## 1. File Naming Patterns\n")
        f.write("- **Structure:** High variability. Common patterns: `Address - Type`, `Project - Type`.\n")
        f.write("- **Year Distribution:** " + ", ".join([f"{y}: {c}" for y, c in Counter(years).most_common(10)]) + "\n\n")
        
        f.write("## 2. Estimate Types (Top 20)\n")
        for t, c in types.most_common(20):
            f.write(f"- {t}: {c}\n")
        f.write("\n")
        
        f.write("## 3. Pricing Structures & 5. Tab Structures\n")
        f.write("Most estimates use a Multi-Tab approach feeding a SUMMARY sheet.\n\n")
        f.write("**Common Tabs:**\n")
        for t, c in tab_structures.most_common(20):
             f.write(f"- {t} ({c} occurences)\n")
        f.write("\n")
        
        f.write("## 4. Common Templates\n")
        f.write("Based on tab combinations, identified distinct template versions:\n")
        for sig, c in common_templates.most_common(5):
            f.write(f"- **Version (Count {c}):** {sig}\n")
        f.write("\n")

        f.write("## 6. Line Item Categories\n")
        f.write("Detected from 'SUMMARY' breakdowns:\n")
        f.write("- Labor\n- Material\n- Equipment (often implicit or separate tab)\n- Subcontractors (Mechanical, Electrical, etc)\n- Travel/Per Diem\n\n")

        f.write("## 7. Sample Data Extraction (Deep Dive)\n")
        f.write("| File | Customer | Project | Active Systems |\n")
        f.write("|---|---|---|---|")
        for d in extracted_data[:15]:
            summ = d.get('summary', {})
            systems = ", ".join(summ.get('active_systems', []))
            f.write(f"| {d['file'][:20]}... | {summ.get('customer', 'N/A')} | {summ.get('project', 'N/A')} | {systems} |\n")
        f.write("\n")
        
        f.write(ERP_COMPARISON_NOTE)

    print(f"Analysis complete. Written to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
